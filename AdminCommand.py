import datetime

from InitBot import (bot, scheduler)
from telebot import types
from InitDB import create_connection
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import asyncio
import os

"""


"""


def SelectActiveTasks(message):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT name_users FROM Telegramm_active_users WHERE Tg_id = %s",(message.chat.id,))
    name_user = cursor.fetchone()
    cursor.execute("SELECT id FROM tasks WHERE who = %s and status = 'В работе'",(name_user[0],))
    AllActiveTasks = cursor.fetchall()
    mark = types.InlineKeyboardMarkup(row_width=3)
    for i in AllActiveTasks:
        mark.add(types.InlineKeyboardButton(f"Номер {i[0]}",
                                            callback_data=f"Номер {i[0]}"))

    mark.add(types.InlineKeyboardButton("Назад",
                                        callback_data="BackAdmin"))
    bot.edit_message_text(text="Выберите, пожалуйста заявку для получения инофрмации: ",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def SelectActiveTasksChange(message):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT name_users FROM Telegramm_active_users WHERE Tg_id = %s",(message.chat.id,))
    name_user = cursor.fetchone()
    cursor.execute("SELECT id FROM tasks WHERE who = %s and status = 'В работе'",(name_user[0],))
    AllActiveTasks = cursor.fetchall()
    mark = types.InlineKeyboardMarkup(row_width=3)
    for i in AllActiveTasks:
        mark.add(types.InlineKeyboardButton(f"Номер {i[0]}",
                                            callback_data=f"Number {i[0]}"))

    mark.add(types.InlineKeyboardButton("Назад",
                                        callback_data="BackAdmin"))
    bot.edit_message_text(text="Выберите, пожалуйста заявку для получения инофрмации: ",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def ChangeInfoAboutAccept(message, id):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT author, city, fio, numberWorker, datePlan, message FROM tasks WHERE id = %s",(id,))
    info = cursor.fetchone()
    mark = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton("Изменить дату принятия", callback_data=f"{id} changeDate")
    button1 = types.InlineKeyboardButton("Назад", callback_data="backToList")
    mark.add(button, button1)
    bot.edit_message_text(f"""Информация по заявке {id}
-----------------
1) Номер: {id}
2) Компания: {info[0]}
3) Автор заявки: {info[2]}
4) Рабочий номер: {info[3]}
5) Дата заявки: {info[4]}
6) Сообщение: {info[5]}
""",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def Change(message, id):
    def times(message):
        dayes = message.text
        bot.send_message(message.chat.id, f"Введите комментарий. Он будет отправлен вам через {dayes} дней")

        bot.register_next_step_handler(message, changes)

    def changes(message):
        time = datetime.datetime.now() + datetime.timedelta(3)
        user_id = id
        scheduler.add_job(ChangeInfo, 'date', run_date=time.strftime("%Y-%m-%d %H:%M:%S"),
                          args=[message, user_id])
        # time.strftime("%Y-%m-%d %H:%M:%S")
        
        bot.send_message(message.chat.id, f"Принятие заявки будет. Я сообщю об этом вам.")

    bot.send_message(message.chat.id, "Введите, через сколько дней принять заявку? (число должно быть целым)")
    bot.register_next_step_handler(message, times)


def ChangeInfo(*args):
    try:
        db = create_connection()
        cursor = db.cursor()
        date = datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("SELECT name_users FROM Telegramm_active_users WHERE Tg_id = %s", (args[0].chat.id,))
        userid = cursor.fetchone()
        cursor.execute("UPDATE tasks SET status = %s , who = %s, created_at = %s, datePlan = %s WHERE id = %s", ('В работе', userid[0],date,date,args[1]))
        mark = types.InlineKeyboardMarkup(row_width=2)
        button1 = types.InlineKeyboardButton(text="👨‍🔧 Список активных заявок 👨‍🔧", callback_data="ActiveTasks")
        button3 = types.InlineKeyboardButton(text="👨‍🔧 Продление заявок 👨‍🔧", callback_data="ProdlTasks")
        button2 = types.InlineKeyboardButton(text="🥇Новые заявки 🥇", callback_data="newTasks")
        button4 = types.InlineKeyboardButton(text="📕 Профиль 📕", callback_data="Profile")
        mark.add(button1, button2, button3, button4)
        bot.send_message(chat_id=args[0].chat.id,
                         text="Уважаемый пользователь, вот команды, которыми вы можете воспользоваться в любое время", 
                         reply_markup=mark,
                         parse_mode="HTML")
        bot.send_message(args[0].chat.id, f"Заявка {args[1]} принята! Комментарий к заявке: {args[0].text}")
        db.commit()

    except:
        bot.send_message(args[0].chat.id, "Не можем принять заявку!")


def SelectNewTasks(message):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT id FROM tasks WHERE status = 'Новая'")
    NewTasks = cursor.fetchall()
    mark = types.InlineKeyboardMarkup(row_width=3)
    for i in NewTasks:
        mark.add(types.InlineKeyboardButton(f"Номер {i[0]}",
                                            callback_data=f"New {i[0]}"))
    mark.add(types.InlineKeyboardButton("Назад",
                                        callback_data="BackAdmin"))

    bot.edit_message_text(text="Выберите, пожалуйста заявку для получения инофрмации: ",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def ChangeInfoNewTask(message, id):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT author, city, fio, numberWorker, datePlan, message FROM tasks WHERE id = %s",(id,))
    info = cursor.fetchone()
    mark = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton("Принять", callback_data=f"{id} AcceptTask")
    button1 = types.InlineKeyboardButton("Назад", callback_data="backToList")
    mark.add(button, button1)
    bot.edit_message_text(f"""Информация по заявке {id}
-----------------
1) Номер: {id}
2) Компания: {info[0]}
3) Автор заявки: {info[2]}
4) Рабочий номер: {info[3]}
5) Дата заявки: {info[4]}
6) Сообщение: {info[5]}
""",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def AcceptNewTask(message, id):
    try:
        db = create_connection()
        cursor = db.cursor()
        date = datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("SELECT name_users FROM Telegramm_active_users WHERE Tg_id = %s", (message.chat.id,))
        userid = cursor.fetchone()
        cursor.execute("UPDATE tasks SET status = %s , who = %s, updated_at = %s WHERE id = %s", ('В работе', userid[0],date,id))
        mark = types.InlineKeyboardMarkup(row_width=2)
        button1 = types.InlineKeyboardButton(text="👨‍🔧 Список активных заявок 👨‍🔧", callback_data="ActiveTasks")
        button3 = types.InlineKeyboardButton(text="👨‍🔧 Продление заявок 👨‍🔧", callback_data="ProdlTasks")
        button2 = types.InlineKeyboardButton(text="🥇Новые заявки 🥇", callback_data="newTasks")
        button4 = types.InlineKeyboardButton(text="📕 Профиль 📕", callback_data="Profile")
        mark.add(button1, button2, button3, button4)
        bot.edit_message_text(chat_id=message.chat.id,
                              text="Уважаемый пользователь, вот команды, которыми вы можете воспользоваться в любое время", 
                              reply_markup=mark,
                              parse_mode="HTML",
                              message_id=message.id)

        bot.send_message(message.chat.id, "Заявка принята!")
        db.commit()

    except:
        bot.send_message(message.chat.id, "Не можем принять заявку!")


def ChangeInformationTask(message, id):
    db = create_connection()
    cursor = db.cursor()
    cursor.execute("SELECT author, city, fio, numberWorker, datePlan, message, updated_at FROM tasks WHERE id = %s",(id,))
    info = cursor.fetchone()
    mark = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton("Закрыть заявку",callback_data=f"{id} CloseTask")
    button1 = types.InlineKeyboardButton("Назад", callback_data = "backToList")
    mark.add(button,button1)
    bot.edit_message_text(f"""Информация по заявке {id}
-----------------
1) Номер: {id}
2) Компания: {info[0]}
3) Автор заявки: {info[2]}
4) Рабочий номер: {info[3]}
5) Дата заявки: {info[4]}
6) Сообщение: {info[5]}
6) Дата принятия заявки: {info[6]}
""",
                          chat_id=message.chat.id,
                          message_id=message.id,
                          reply_markup=mark)
    db.commit()


def CloseTask(message, id):
    def inputComment(message):
        db = create_connection()
        cursor = db.cursor()
        date = datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        try:
            cursor.execute("UPDATE tasks SET status = 'Выполнена', dateFact = %s, comment = %s where id = %s;",(date,message.text,id))
            bot.send_message(message.chat.id,
                             "Заявка закрыта!")        
        except:
            bot.send_message(message.chat.id,
                             "Заявка не закрыта. Повторите попытку!")
        db.commit()
    bot.send_message(chat_id=message.chat.id,
                     text="Введите комментарий по заявке: ")

    bot.register_next_step_handler(message, inputComment)


def ExportInfoAboutWork(message):
    db = create_connection()
    cursor = db.cursor()

    cursor.execute("SELECT name_users FROM Telegramm_active_users WHERE Tg_id = %s", (message.chat.id,))
    name_worker = cursor.fetchone()[0]

    cursor.execute("SELECT id, message, fio, numberWorker,author, status, datePlan, dateFact FROM tasks WHERE who = %s",(name_worker,))
    all_works = cursor.fetchall()

    wb = Workbook() 
    ws = wb.active
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['G'].width = 48
    ws.column_dimensions['C'].width = 25
    thin = Side(border_style="thin", color="000000")
    ws.title = f"{name_worker}"
    """
    Все заявки
    """
    company = []
    SredTime = []
    for i in range(len(all_works)):
        ws.row_dimensions[i+1].height = 45
        ws[f'A{i+2}'].alignment = Alignment(wrap_text=True)
        ws[f'A{i+2}'].fill = PatternFill('solid', fgColor="fff2f2f2")
        if all_works[i][7] is None or all_works[i][6] is None:
            count_day = " - "
            SredTime.append(0)
        else:
            count_day = abs((all_works[i][7] - all_works[i][6]).days)
            SredTime.append(count_day)
            company.append(all_works[i][4])
        ws[f'A{i+2}'] = f"""Информация по заявке {all_works[i][0]}
-----------------
1) Номер: {all_works[i][0]}
2) Компания: {all_works[i][4]}
3) Автор заявки: {all_works[i][2]}
4) Рабочий номер: {all_works[i][3]}
5) Сообщение: {all_works[i][1]}
6) Время выполняния: {count_day} дней
7) Статус: {all_works[i][5]}
"""
        ws[f'A{i+2}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    """
    Статистика:
    """
    ws['C2'] = "Общее количество заявок: "
    ws['C2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['C2'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['C2'].fill = PatternFill('solid', fgColor= "ffd8e4bc")
    ws['D2'] = len(all_works)
    ws['D2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['D2'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['D2'].fill = PatternFill('solid', fgColor= "ffc4d79b")

    ws['C4'] = "Среднее время выполнения 1 заявки (в днях): "
    ws['C4'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['C4'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['C4'].fill = PatternFill('solid', fgColor= "ffb8cce4")
    ws['D4'] = sum(SredTime)//len(SredTime)
    ws['D4'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['D4'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['C4'].fill = PatternFill('solid', fgColor= "ff95b3d7")

    name_company = set(company)
    g = 3
    ws['G2'] = "Компания"
    ws['G2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['G2'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['G2'].fill = PatternFill('solid', fgColor= "ffe6b7b8")
    ws['H2'] = "Количество"
    ws.column_dimensions['H'].width = 10.89
    ws['H2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['H2'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws['H2'].fill = PatternFill('solid', fgColor= "ffe6b7b8")
    for i in name_company:
        ws[f'G{g}'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws[f'H{g}'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws[f'G{g}'].fill = PatternFill('solid', fgColor= "fffcf6f6")
        ws[f'H{g}'].fill = PatternFill('solid', fgColor= "fffcf6f6")
        ws[f'G{g}'] = i
        ws[f'H{g}'] = company.count(i)
        ws[f'G{g}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws[f'H{g}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        g += 1

    wb.save(f"{name_worker}.xlsx")
    f = open(f"{name_worker}.xlsx", "rb")
    bot.send_document(message.chat.id,
                      document=f,
                      caption=f"Статистика работника {name_worker}")
    f.close()
    os.remove(f'{name_worker}.xlsx')


async def start_schedule():
    while True:
        await asyncio.sleep(10)


if __name__ == "__main__":
    Change("text", 1)
